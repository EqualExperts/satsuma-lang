#!/usr/bin/env python3
"""Generate the Phase 0.5 probe spreadsheet arms.

Two workbooks are produced from the same hand-authored intent, so the arms
are paired to each other (the best control available for ~$8, per the PRD):

  - X-P0: tidy. One tab per mapping, fixed columns, plain header row, one row
    per field-level arrow. Adversarially favourable to Excel — this is the
    headline pair with S+.
  - X-P2: realistic-messy. P0 plus the messiness primitives already proven in
    archive/features/04-excel-to-stm-skill/test-data/generate_test_spreadsheets.py
    — a free-text Notes column, merged header cells, semantics in cell fill
    colour with NO legend, a stale "Archived" tab, and a multi-row title block
    that pushes the headers off row 1. P2 is where fill colour (invisible to
    pandas.read_excel) carries meaning a human sees and an agent misses.

Run: python3 generate_probe_spreadsheets.py
Writes the two .xlsx files next to this script.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


# ── Styling primitives (reused from the archived generator) ────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", wrap_text=True)


def style_header_row(ws, row, ncols, fill_color="1F4E79", font_color="FFFFFF"):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color=font_color, size=11)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


# ── The shared mapping data (the one intent, as flat rows) ─────────────────
# Each mapping is a list of rows: source, source_type, target, target_type,
# transform, required, notes.

MAPPING_COLUMNS = [
    "Source Field", "Source Type",
    "Target Field", "Target Type",
    "Transformation", "Required?", "Notes",
]


def claim_normalisation_rows():
    return [
        ("claim_header.claim_id", "VARCHAR(20)", "claim_key", "VARCHAR(20)", "Direct copy", "Y", "PK"),
        ("claim_header.policy_no", "VARCHAR(30)", "policy_ref", "VARCHAR(30)", "Direct copy", "N", "indexed"),
        ("claim_header.claim_type", "VARCHAR(15)", "claim_type_code", "VARCHAR(2)",
         "Map: auto->AU, home->HO, life->LI, health->HE", "N", ""),
        ("claim_header.reported_at", "TIMESTAMPTZ", "reported_date", "TIMESTAMP_NTZ", "Convert to UTC", "N", ""),
        ("claim_header.loss_amount", "DECIMAL(14,2)", "loss_usd", "DECIMAL(14,2)",
         "Convert to USD using fx_rates lookup on claim_header.currency, then round 2", "Y", ""),
        ("claim_header.loss_amount", "DECIMAL(14,2)", "loss_source", "VARCHAR(10)",
         "coalesce 0", "N", ""),
        ("claim_header.vehicles", "list_of record", "vehicle_count", "INT", "count", "N", ""),
        ("claim_header.parties", "list_of record", "party_count", "INT", "count", "N", ""),
        ("claim_header.vehicles.damage_extent", "VARCHAR(15)", "max_damage", "VARCHAR(15)",
         "Pick the worst damage across all vehicles on the claim", "N", "enum: none,minor,moderate,severe,total,scratch"),
        ("claim_header.vehicles.photos", "list_of record", "photos", "list_of record", "flatten", "N", ""),
        ("vehicles.photos.photo_id", "VARCHAR(36)", "photos.photo_ref", "VARCHAR(36)", "trim", "Y", "inside flatten"),
        ("vehicles.photos.angle", "VARCHAR(10)", "photos.view", "VARCHAR(10)", "lowercase", "N", "inside flatten"),
        ("claim_header.adjuster_id", "VARCHAR(20)", "adjuster_ref", "VARCHAR(20)", "trim", "N", ""),
        ("(computed — no source)", "", "is_open", "BOOLEAN",
         "True if claim_header.status is 'open' or 'under_review', false otherwise.", "N", ""),
    ]


def party_extract_rows():
    return [
        ("claim_header.claim_id", "VARCHAR(20)", "claim_key", "VARCHAR(20)", "Direct copy", "Y", ""),
        ("parties[].party_role", "VARCHAR(20)", "rows[].role", "VARCHAR(20)", "trim | uppercase", "Y", "each parties -> rows"),
        ("parties[].name", "VARCHAR(120)", "rows[].display_name", "VARCHAR(120)", "trim", "N", ""),
        ("parties[].contact_phone", "VARCHAR(20)", "rows[].phone_e164", "VARCHAR(20)",
         "Format to E.164, assuming US country code if no + prefix", "N", "PII"),
    ]


def vehicle_extract_rows():
    return [
        ("claim_header.claim_id", "VARCHAR(20)", "claim_key", "VARCHAR(20)", "Direct copy", "Y", ""),
        ("vehicles[].vin", "VARCHAR(17)", "rows[].vin", "VARCHAR(17)", "uppercase", "Y", "each vehicles -> rows"),
        ("vehicles[].make", "VARCHAR(40)", "rows[].description", "VARCHAR(100)",
         "vehicles.make || ' ' || vehicles.model || ' ' || vehicles.year", "N", ""),
        ("vehicles[].damage_extent", "VARCHAR(15)", "rows[].damage_class", "VARCHAR(15)",
         "Map: none->N, minor->M, moderate->M, severe->S, total->T (scratch unhandled)", "N", ""),
        ("vehicles[].estimate", "DECIMAL(12,2)", "rows[].estimate_usd", "DECIMAL(12,2)", "round 2", "N", ""),
    ]


def status_snapshot_rows():
    return [
        ("claim_fact.claim_key", "VARCHAR(20)", "claim_key", "VARCHAR(20)", "Direct copy", "Y", "PK, ref claim_fact.claim_key"),
        ("claim_fact.is_open", "BOOLEAN", "open_flag", "BOOLEAN", "Direct copy", "N", ""),
        ("claim_fact.loss_usd", "DECIMAL(14,2)", "total_exposure", "DECIMAL(14,2)",
         "Sum exposure across all open claims for the same policy", "N", ""),
    ]


def payment_extract_rows():
    return [
        ("claim_fact.claim_key", "VARCHAR(20)", "claim_key", "VARCHAR(20)", "Direct copy", "Y", "ref claim_fact.claim_key"),
        ("claim_fact.loss_usd", "DECIMAL(14,2)", "paid_amount", "DECIMAL(14,2)", "round 2", "N", ""),
        ("claim_fact.reported_date", "TIMESTAMP_NTZ", "paid_at", "TIMESTAMP_NTZ", "to_utc", "N", ""),
    ]


def fraud_assessment_rows():
    return [
        ("claim_fact.claim_key", "VARCHAR(20)", "claim_key", "VARCHAR(20)", "Direct copy", "Y", "PK, ref claim_fact.claim_key"),
        ("claim_fact.party_count", "INT", "risk_score", "INT", "multiply 10", "N", ""),
        ("claim_fact.party_count", "INT", "is_flagged", "BOOLEAN",
         "True if party_count > 7, false otherwise.", "N", ""),
    ]


MAPPINGS = [
    ("claim_normalisation", claim_normalisation_rows()),
    ("party_extract", party_extract_rows()),
    ("vehicle_extract", vehicle_extract_rows()),
    ("status_snapshot", status_snapshot_rows()),
    ("payment_extract", payment_extract_rows()),
    ("fraud_assessment", fraud_assessment_rows()),
]


# ── Schema declarations (so the Excel arms carry what .stm and .md carry) ──
# The PRD's totality control exists to stop one arm saying less than another.
# The first version of this generator shipped no schema tabs at all, so the
# Excel arms could not see types, enums, pk/required/pii metadata, or unmapped
# fields — confounding the primary statistic and T5 quality. These tabs are
# the fix. Each schema is one row per field, with the enum stated verbatim.

SCHEMA_COLUMNS = ["Field", "Type", "Required", "Notes"]


def _schema_rows():
    return {
        "claim_header (source)": [
            ("claim_id", "VARCHAR(20)", "Y", "PK"),
            ("policy_no", "VARCHAR(30)", "Y", ""),
            ("claim_type", "VARCHAR(15)", "Y", "enum: auto, home, life, health"),
            ("reported_at", "TIMESTAMPTZ", "Y", ""),
            ("incident_at", "TIMESTAMPTZ", "N", ""),
            ("loss_amount", "DECIMAL(14,2)", "N", ""),
            ("currency", "CHAR(3)", "N", "default USD"),
            ("adjuster_id", "VARCHAR(20)", "N", ""),
            ("status", "VARCHAR(20)", "Y", ""),
            ("vehicles", "list_of record", "N", "nested"),
            ("vehicles[].vin", "VARCHAR(17)", "Y", ""),
            ("vehicles[].make", "VARCHAR(40)", "N", ""),
            ("vehicles[].model", "VARCHAR(40)", "N", ""),
            ("vehicles[].year", "INT", "N", ""),
            ("vehicles[].damage_extent", "VARCHAR(15)", "N",
             "enum: none, minor, moderate, severe, total, scratch"),
            ("vehicles[].estimate", "DECIMAL(12,2)", "N", ""),
            ("vehicles[].photos", "list_of record", "N", "nested"),
            ("vehicles[].photos[].photo_id", "VARCHAR(36)", "Y", ""),
            ("vehicles[].photos[].angle", "VARCHAR(10)", "N",
             "enum: front, rear, left, right, interior"),
            ("parties", "list_of record", "N", "nested"),
            ("parties[].party_role", "VARCHAR(20)", "Y", ""),
            ("parties[].name", "VARCHAR(120)", "N", ""),
            ("parties[].contact_phone", "VARCHAR(20)", "N", "PII"),
        ],
        "policy_dim (lookup)": [
            ("policy_no", "VARCHAR(30)", "Y", "PK"),
            ("product_code", "VARCHAR(10)", "Y", ""),
            ("policyholder_id", "VARCHAR(20)", "Y", ""),
            ("effective_date", "DATE", "N", ""),
            ("expiry_date", "DATE", "N", ""),
            ("territory", "VARCHAR(10)", "N", ""),
        ],
        "fx_rates (lookup)": [
            ("currency", "CHAR(3)", "Y", "PK"),
            ("rate_to_usd", "DECIMAL(10,6)", "Y", ""),
        ],
        "claim_fact (target)": [
            ("claim_key", "VARCHAR(20)", "N", "PK"),
            ("policy_ref", "VARCHAR(30)", "N", "indexed"),
            ("claim_type_code", "VARCHAR(2)", "N", ""),
            ("reported_date", "TIMESTAMP_NTZ", "N", ""),
            ("loss_usd", "DECIMAL(14,2)", "Y", ""),
            ("loss_source", "VARCHAR(10)", "N", ""),
            ("vehicle_count", "INT", "N", ""),
            ("party_count", "INT", "N", ""),
            ("max_damage", "VARCHAR(15)", "N", ""),
            ("photos", "list_of record", "N", "nested"),
            ("photos[].photo_ref", "VARCHAR(36)", "Y", ""),
            ("photos[].view", "VARCHAR(10)", "N", ""),
            ("adjuster_ref", "VARCHAR(20)", "N", ""),
            ("is_open", "BOOLEAN", "N", ""),
        ],
        "party_dim (target)": [
            ("claim_key", "VARCHAR(20)", "Y", ""),
            ("rows", "list_of record", "N", "nested"),
            ("rows[].role", "VARCHAR(20)", "Y", ""),
            ("rows[].display_name", "VARCHAR(120)", "N", ""),
            ("rows[].phone_e164", "VARCHAR(20)", "N", ""),
        ],
        "vehicle_dim (target)": [
            ("claim_key", "VARCHAR(20)", "Y", ""),
            ("rows", "list_of record", "N", "nested"),
            ("rows[].vin", "VARCHAR(17)", "Y", ""),
            ("rows[].description", "VARCHAR(100)", "N", ""),
            ("rows[].damage_class", "VARCHAR(15)", "N", ""),
            ("rows[].estimate_usd", "DECIMAL(12,2)", "N", ""),
        ],
        "claim_status_snapshot (target)": [
            ("claim_key", "VARCHAR(20)", "N", "PK, ref claim_fact.claim_key"),
            ("open_flag", "BOOLEAN", "N", ""),
            ("total_exposure", "DECIMAL(14,2)", "N", ""),
        ],
        "payment_fact (target)": [
            ("payment_id", "VARCHAR(36)", "N", "PK"),
            ("claim_key", "VARCHAR(20)", "N", "ref claim_fact.claim_key"),
            ("paid_amount", "DECIMAL(14,2)", "N", ""),
            ("paid_at", "TIMESTAMP_NTZ", "N", ""),
        ],
        "fraud_flag (target)": [
            ("claim_key", "VARCHAR(20)", "N", "PK, ref claim_fact.claim_key"),
            ("risk_score", "INT", "N", ""),
            ("is_flagged", "BOOLEAN", "N", ""),
        ],
    }


def add_schema_tabs(wb):
    """Add one tab per schema, before the mapping tabs, to either workbook."""
    schemas = _schema_rows()
    # Insert after the cover/README tab (index 1) so schemas precede mappings.
    insert_at = 1 if wb.sheetnames and wb.sheetnames[0] == "README" else 0
    for offset, (name, rows) in enumerate(schemas.items()):
        ws = wb.create_sheet(name, index=insert_at + offset)
        ws.append(SCHEMA_COLUMNS)
        style_header_row(ws, 1, len(SCHEMA_COLUMNS))
        for r in rows:
            ws.append(r)
        auto_width(ws)


# ── X-P0: tidy ─────────────────────────────────────────────────────────────

def build_p0():
    wb = openpyxl.Workbook()
    # A cover/README tab first (P0 still has one — it's tidy, not bare).
    cover = wb.active
    cover.title = "README"
    cover["A1"] = "Meridian Mutual — Claims Mapping (P0 tidy)"
    cover["A1"].font = Font(bold=True, size=14)
    cover["A3"] = "One tab per mapping. Columns: " + ", ".join(MAPPING_COLUMNS)
    cover.column_dimensions["A"].width = 90

    for name, rows in MAPPINGS:
        ws = wb.create_sheet(name)
        ws.append(MAPPING_COLUMNS)
        style_header_row(ws, 1, len(MAPPING_COLUMNS))
        for r in rows:
            ws.append(r)
        auto_width(ws)

    add_schema_tabs(wb)

    path = os.path.join(SCRIPT_DIR, "meridian-claims-P0.xlsx")
    wb.save(path)
    print(f"Created: {path}")


# ── X-P2: realistic-messy ──────────────────────────────────────────────────
# P0 + free-text Notes, merged headers, fill-colour-as-semantics (NO legend),
# multi-row title block pushing headers off row 1, and a stale "Archived" tab.

P2_NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")  # "needs review" yellow
P2_PII_FILL = PatternFill("solid", fgColor="FCE4D6")   # peach = PII
P2_COMPUTED_FILL = PatternFill("solid", fgColor="E2EFDA")  # green = computed/no-source
# The load-bearing P2 hazard: this fill carries meaning pandas cannot read.
P2_AMBIGUITY_FILL = PatternFill("solid", fgColor="D9E1F2")  # blue = ambiguous transform


def build_p2():
    wb = openpyxl.Workbook()
    # Claim the default empty Sheet so P2 ships no stray blank tab, and use it
    # as the cover/README (P0 has one; P2 must too, or the arms are not paired).
    cover = wb.active
    cover.title = "README"
    cover["A1"] = "Meridian Mutual — Claims Mapping (P2 working draft)"
    cover["A1"].font = Font(bold=True, size=14)
    cover["A3"] = "Tabs: schemas first, then mappings. Colour coding is intentional; no legend is provided."
    cover.column_dimensions["A"].width = 90

    add_schema_tabs(wb)

    # Tab: a multi-row title block — headers are NOT on row 1.
    ws = wb.create_sheet("claim_normalisation")
    ws["A1"] = "Meridian Mutual"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = "Claims Mapping — v3 (working draft)"
    ws["A2"].font = Font(italic=True, size=11)
    ws["A3"] = "Owner: Data Eng | Last updated: see Changelog tab"
    # Merged header cell spanning the column set on row 5 (off row 1).
    ws.merge_cells("A5:G5")
    ws["A5"] = "FIELD MAPPING"
    ws["A5"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A5"].fill = HEADER_FILL
    ws["A5"].alignment = Alignment(horizontal="center")
    # Actual column headers on row 6.
    for col, h in enumerate(MAPPING_COLUMNS, 1):
        ws.cell(row=6, column=col, value=h)
    style_header_row(ws, 6, len(MAPPING_COLUMNS))

    rows = MAPPINGS[0][1]
    for i, r in enumerate(rows):
        excel_row = 7 + i
        for col, val in enumerate(r, 1):
            ws.cell(row=excel_row, column=col, value=val)
        # Fill-colour semantics (NO legend anywhere in the workbook).
        target_field = r[2]
        transform = r[4]
        notes = r[6]
        if target_field == "is_open":
            for c in range(1, 8):
                ws.cell(row=excel_row, column=c).fill = P2_COMPUTED_FILL
        elif target_field == "loss_usd" or target_field == "total_exposure":
            # The underspecified-rounding ambiguities — marked by fill only.
            for c in range(1, 8):
                ws.cell(row=excel_row, column=c).fill = P2_AMBIGUITY_FILL
        elif "PII" in notes or "contact_phone" in (r[0] or ""):
            for c in range(1, 8):
                ws.cell(row=excel_row, column=c).fill = P2_PII_FILL
        elif notes:
            for c in range(1, 8):
                ws.cell(row=excel_row, column=c).fill = P2_NOTE_FILL

    auto_width(ws)

    # Remaining mappings: standard header row but with a free-text Notes column
    # that mixes rules and commentary (P1-style, but P2 keeps it).
    for name, rows in MAPPINGS[1:]:
        ws2 = wb.create_sheet(name)
        ws2.append(MAPPING_COLUMNS)
        style_header_row(ws2, 1, len(MAPPING_COLUMNS))
        for r in rows:
            ws2.append(r)
        # P2 fill semantics, keyed on the full target path so rows[]-prefixed
        # fields match. Only the planted ambiguities and PII are coloured —
        # marking unambiguous fields (e.g. estimate_usd) would leak a wrong
        # answer on T5.
        P2_ROUND_TARGETS = {"total_exposure", "paid_amount"}  # A1: underspecified rounding
        P2_PHONE_TARGET = "rows[].phone_e164"                  # A4: implicit country
        P2_PII_SOURCE = "parties[].contact_phone"
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
            src = str(row[0].value or "")
            tgt = str(row[2].value or "")
            notes = str(row[6].value or "")
            if src == P2_PII_SOURCE or "PII" in notes:
                for cell in row:
                    cell.fill = P2_PII_FILL
            elif tgt == P2_PHONE_TARGET:
                for cell in row:
                    cell.fill = P2_AMBIGUITY_FILL
            elif tgt in P2_ROUND_TARGETS:
                for cell in row:
                    cell.fill = P2_AMBIGUITY_FILL
        auto_width(ws2)

    # A stale "Archived" tab that a tidy author would have deleted.
    ws_arch = wb.create_sheet("Archived v2 mapping")
    ws_arch["A1"] = "Superseded by claim_normalisation tab above. Do not use."
    ws_arch["A1"].font = Font(strikethrough=True, color="999999")
    ws_arch.append([])
    ws_arch.append(MAPPING_COLUMNS)
    style_header_row(ws_arch, 3, len(MAPPING_COLUMNS))
    ws_arch.append(("claim_header.claim_id", "VARCHAR(20)", "claim_id", "VARCHAR(20)",
                    "Direct copy", "Y", "OLD — renamed to claim_key in v3"))
    auto_width(ws_arch)

    # A Changelog tab (matches the multi-tab fixture style).
    ws_log = wb.create_sheet("Changelog")
    ws_log.append(["Date", "Author", "Change"])
    style_header_row(ws_log, 1, 3, fill_color="808080")
    ws_log.append(("2025-11-02", "R. Varga", "Initial draft"))
    ws_log.append(("2025-11-20", "R. Varga", "Added party_extract and vehicle_extract tabs"))
    ws_log.append(("2025-12-01", "L. Okafor", "Marked loss_usd rounding as needing review (see fill)"))
    auto_width(ws_log)

    path = os.path.join(SCRIPT_DIR, "meridian-claims-P2.xlsx")
    wb.save(path)
    print(f"Created: {path}")


if __name__ == "__main__":
    build_p0()
    build_p2()
    print("\nDone — 2 probe workbooks created.")
